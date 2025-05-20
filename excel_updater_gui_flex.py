import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
import os

# Reference map definition
raw_reference_map = {
    "international freight, Freight Costs": "51300",
    "customs clearance & admin, Customs Clearance and Admin": "51400",
    "isf fee": "51500",
    "freight insurance": "51000",
    "material": "50000",
    "others, others_round up": "59000",
    "destination pierpass, Destination Pier Pass": "59110",
    "drayage": "59120",
    "exam": "59130",
    "detention": "59140",
    "chassis, Destination Chassis Fee": "59150",
    "dry run": "59160",
    "storage": "59170",
    "demurrage, destination demurrage": "59180",
    "per diem": "59190",
    "terminal fee": "59200",
    "handling fees, Handling Fee": "59210",
    "ams": "59220",
    "pre pull": "59230",
    "duty, custom duty 7501": "59240",
    "exwork": "59250",
    "warehouse in/out": "59260"
}

# Flatten the map for partial match lookup
reference_map = {}
for key_string, value in raw_reference_map.items():
    keys = [k.strip().lower() for k in key_string.split(",")]
    for key in keys:
        reference_map[key] = value

# Function to update the Excel file
def update_excel(file_path):
    try:
        wb = load_workbook(file_path)
        ws = wb.active

        for row in ws.iter_rows(min_row=2):
            col_g = row[6]  # Column G
            col_e = row[4]  # Column E

            if col_g.value:
                cleaned_text = str(col_g.value).strip().lower()
                for keyword, code in reference_map.items():
                    if keyword in cleaned_text:
                        col_e.value = code
                        break  # stop after first match

        folder, original_file = os.path.split(file_path)
        name, ext = os.path.splitext(original_file)
        new_filename = f"{name}_updatedfortrader{ext}"
        new_file_path = os.path.join(folder, new_filename)
        wb.save(new_file_path)

        messagebox.showinfo("Success", f"✅ Updated File saved to:\n{new_file_path}")
    except Exception as e:
        messagebox.showerror("Error", f"❌ Failed to update file:\n{str(e)}")

# GUI Setup
def browse_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if file_path:
        entry_file_path.delete(0, tk.END)
        entry_file_path.insert(0, file_path)

def run_update():
    path = entry_file_path.get()
    if not os.path.isfile(path):
        messagebox.showwarning("Warning", "Please select a valid file.")
        return
    update_excel(path)

# Default file path
default_path = r"C:\Users\eric.gao\Downloads\lastest bill.xlsx"

root = tk.Tk()
root.title("Excel Bill Updater")
root.geometry("500x150")
root.resizable(False, False)

# File path input
tk.Label(root, text="Excel File Path:").pack(pady=(10, 0))
entry_file_path = tk.Entry(root, width=60)
entry_file_path.pack(pady=5)
entry_file_path.insert(0, default_path)  # Set default path
tk.Button(root, text="Browse...", command=browse_file).pack()

# Run button
tk.Button(root, text="Run Update", command=run_update, bg="#4CAF50", fg="white", height=2).pack(pady=10)

root.mainloop()
