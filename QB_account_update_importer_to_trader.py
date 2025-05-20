from openpyxl import load_workbook
import os

# Path to your file (update filename as needed)
file_path = r"C:\Users\eric.gao\Downloads\lastest bill.xlsx"

# Original-style input: comma-separated keys
raw_reference_map = {
    "international freight": "51300",
    "customs clearance & admin": "51400",
    "isf fee": "51500",
    "freight insurance": "51000",
    "material": "50000",
    "others, others_round up": "59000",
    "destination pierpass, Destination Pier Pass": "59110",
    "drayage": "59120",
    "exam": "59130",
    "detention": "59140",
    "chassis, Destination Chassis Fee": "59150",
    "dry run": "59160",
    "storage": "59170",
    "demurrage, destination demurrage": "59180",
    "per diem": "59190",
    "terminal fee": "59200",
    "handling fees, Handling Fee": "59210",
    "ams": "59220",
    "pre pull": "59230",
    "duty, custom duty 7501": "59240",
    "exwork": "59250",
    "warehouse in/out": "59260",
    "bonds": "59270"
}

# Flatten the map into individual keyword: value pairs
reference_map = {}
for key_string, value in raw_reference_map.items():
    keys = [k.strip().lower() for k in key_string.split(",")]
    for key in keys:
        reference_map[key] = value

# Load the workbook
wb = load_workbook(file_path)
ws = wb.active  # or use wb['SheetName'] if needed

# Loop through rows (starting from row 2 to skip headers)
for row in ws.iter_rows(min_row=2):
    col_g = row[6]  # Column G
    col_e = row[4]  # Column E

    if col_g.value:
        cleaned_text = str(col_g.value).strip().lower()
        if cleaned_text in reference_map:
            col_e.value = reference_map[cleaned_text]

# Save the updated file
folder, original_file = os.path.split(file_path)
name, ext = os.path.splitext(original_file)
new_filename = f"{name}_updatedfortrader{ext}"
new_file_path = os.path.join(folder, new_filename)

wb.save(new_file_path)
print(f"✅ Updated file saved as: {new_file_path}")
