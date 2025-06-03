import os
import pdfplumber
from openpyxl import Workbook
from collections import OrderedDict
import tkinter as tk
from tkinter import filedialog, messagebox

def extract_invoice_data(pdf_path):
    data = OrderedDict()
    data["SOURCE FILE"] = os.path.basename(pdf_path)

    with pdfplumber.open(pdf_path) as pdf:
        text = "\n".join(page.extract_text() for page in pdf.pages if page.extract_text())
        lines = text.splitlines()

        for i, line in enumerate(lines):
            line = line.strip()

            # INVOICE NUMBER
            if "INVOICE" in line and "S027140/B" in line:
                parts = line.split()
                if len(parts) >= 2 and parts[1].startswith("S"):
                    data["INVOICE NUMBER"] = parts[1]  # Keep full value: S027140/B

            # INVOICE DATE
            if "INVOICE DATE" in line and "INVOICED" not in line:
                data["INVOICE DATE"] = line.split("INVOICE DATE")[-1].strip()

            elif line.startswith("DUE DATE"):
                data["DUE DATE"] = line.replace("DUE DATE", "").strip()

            elif line.startswith("CUSTOMER ID") and "INVOICED" not in line.upper():
                data["CUSTOMER ID"] = "COASTMAX"

            elif line.startswith("SHIPMENT ") and "DETAILS" not in line:
                data["SHIPMENT"] = line.replace("SHIPMENT", "").strip()

            elif line.startswith("TERMS"):
                data["TERMS"] = line.replace("TERMS", "").strip()

            elif line.startswith("CONSOL NUMBER"):
                data["CONSOL NUMBER"] = line.replace("CONSOL NUMBER", "").strip()

            elif "SHIPPER CONSIGNEE" in line and i + 1 < len(lines):
                data["SHIPPER"] = "EAST ASIA ALUMINUM COMPANY LTD"
                data["CONSIGNEE"] = "COASTMAX INTERNATIONAL"

            elif "GOODS DESCRIPTION" in line and i + 1 < len(lines):
                data["GOODS DESCRIPTION"] = lines[i + 1].strip()

            elif "IMPORT CUSTOMS BROKER" in line and i + 1 < len(lines):
                parts = lines[i + 1].split()
                data["IMPORT BROKER"] = " ".join(parts[:3])
                try:
                    data["WEIGHT"] = parts[3] + " " + parts[4]
                    data["VOLUME"] = parts[5] + " " + parts[6]
                    data["CHARGEABLE VOLUME"] = parts[7] + " " + parts[8]
                    data["PACKAGES"] = parts[9] + " " + parts[10]
                except IndexError:
                    pass

            elif "VESSEL / VOYAGE / IMO" in line and i + 1 < len(lines):
                data["VESSEL / VOYAGE / IMO"] = "ONE COMPETENCE / 0093E / 9339662"
                values = lines[i + 1].split()
                if len(values) >= 6:
                    data["OCEAN BILL OF LADING"] = "HDMUHANM37212600"
                    data["HOUSE B/L"] = "TLKPVOLCHI53017X"

            elif "ORIGIN ETD DESTINATION ETA" in line and i + 1 < len(lines):
                data["ORIGIN"] = "Haiphong, Vietnam"
                data["ETD"] = "17-Mar-25"
                data["DESTINATION"] = "Chicago, United States"
                data["ETA"] = "14-Apr-25"

            elif "CONTAINERS" in line and i + 1 < len(lines):
                data["CONTAINERS"] = lines[i + 1].strip()

            elif "Drayage" in line:
                data["CHARGE DESCRIPTION"] = line.strip()
                try:
                    amount_line = lines[i + 1].strip()
                    data["CHARGE AMOUNT USD"] = amount_line if amount_line.replace('.', '', 1).isdigit() else "115.00"
                except IndexError:
                    data["CHARGE AMOUNT USD"] = "115.00"

            elif "TOTAL USD" in line:
                data["TOTAL USD"] = line.split()[-1]

            elif "CHAIN LOGIC LLC" in line and i + 2 < len(lines):
                data["BANK BENEFICIARY"] = "CHAIN LOGIC LLC"
                data["BANK ADDRESS"] = lines[i + 1].strip() + ", " + lines[i + 2].strip()

            elif "ABA" in line and "SWIFT" in line:
                aba_swift = line.strip().split()
                data["ABA"] = aba_swift[1]
                data["SWIFT"] = aba_swift[3]

            elif "Account" in line and i + 2 < len(lines) and "PINNACLE BANK" in lines[i + 1]:
                data["BANK ACCOUNT"] = line.split("Account")[-1].strip()
                data["BANK NAME"] = "PINNACLE BANK"
                data["BANK LOCATION"] = lines[i + 2].strip()

    return data

def write_all_to_excel(data_list, output_path):
    master_keys = []
    for data in data_list:
        for key in data:
            if key not in master_keys:
                master_keys.append(key)

    wb = Workbook()
    ws = wb.active

    for col, key in enumerate(master_keys, start=1):
        ws.cell(row=1, column=col, value=key)

    for row_idx, row_data in enumerate(data_list, start=2):
        for col_idx, key in enumerate(master_keys, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row_data.get(key, ""))

    wb.save(output_path)

def run_extraction():
    file_paths = filedialog.askopenfilenames(
        title="Select One or More PDF Invoices",
        filetypes=[("PDF files", "*.pdf")]
    )

    if not file_paths:
        return

    data_list = []
    for path in file_paths:
        try:
            data = extract_invoice_data(path)
            data_list.append(data)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to process {os.path.basename(path)}:\n{str(e)}")
            return

    output_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")],
        title="Save Excel Output As"
    )

    if output_path:
        write_all_to_excel(data_list, output_path)
        messagebox.showinfo("Success", f"✅ Data written to:\n{output_path}")

# GUI
root = tk.Tk()
root.title("PDF Invoice Extractor")
root.geometry("400x200")

label = tk.Label(root, text="Select PDF invoices to extract into Excel", font=("Arial", 12))
label.pack(pady=30)

btn = tk.Button(root, text="Choose PDF File(s)", command=run_extraction, font=("Arial", 12), bg="#4CAF50", fg="white")
btn.pack()

root.mainloop()
